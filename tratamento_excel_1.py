import openpyxl as xl
from openpyxl.styles import NamedStyle
from datetime import date
import datetime
import pandas as pd
import timeit

 
# Abrindo o arquivo fonte - sai do site do Linx UX
tic=timeit.default_timer()

pivot = xl.load_workbook(r"diretoriodoarquivobase")
vendas = pivot.worksheets[0]


# Abrindo o arquivo que vai pro DW

ticket_ux = xl.load_workbook(r"diretoriodoarquivoDW")
dw = ticket_ux.active

#Limpando o arquivo (exceto os títulos das colunas)

for row in dw['A2:W105000']:
    for cell in row:
        cell.value = None

#Definindo número de linhas e colunas

linha = vendas.max_row
col = vendas.max_column

# Copiando as células do arquivo do Linx UX para o do DW

for i in range (2, linha + 1):
    for j in range (1, col + 1):
        c = vendas.cell(row = i, column = j)
        dw.cell(row = i, column = j).value = c.value   
    
for i in range(2, linha + 1):
    c = dw.cell(row = i, column = 1)
    c.value = float(c.value)
    
for i in range(2, linha + 1):
    c = dw.cell(row = i, column = 3)
    c.value = float(c.value)

for i in range(2, linha + 1):
    c = dw.cell(row = i, column = 4)
    c.number_format = 'dd/mmm/yy'
    
# Salvando o arquivo do DW

ticket_ux.save(filename = r"diretorioarquivofinal")

toc=timeit.default_timer()

tempototal=toc - tic
print(tempototal)
